import openpyxl, zipfile, re, json, os

SRC = r"C:/Users/Administrator/Desktop/收集表.xlsx"
PROJ = r"C:/Users/Administrator/WorkBuddy/2026-07-31-16-56-40/ai-hairstyle-ar"
IMG_DIR = os.path.join(PROJ, "public/images/hair")
os.makedirs(IMG_DIR, exist_ok=True)

# 1) cellimages.xml: ID(name attr) -> rId ; rels: rId -> media file
z = zipfile.ZipFile(SRC)
cellimages = z.read('xl/cellimages.xml').decode('utf-8', 'ignore')
rels = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8', 'ignore')
blocks = re.findall(r'<etc:cellImage>.*?</etc:cellImage>', cellimages, re.S)
id2rid = {}
for b in blocks:
    m = re.search(r'name="(ID_[0-9A-Fa-f]{32})"', b)
    r = re.search(r'r:embed="(rId\d+)"', b)
    if m and r:
        id2rid[m.group(1)] = r.group(1)
rid2media = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
id2media = {i: rid2media.get(rid) for i, rid in id2rid.items()}
print("id2rid:", len(id2rid), "| rid2media:", len(rid2media))

# 2) read sheet rows
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
FACE_WORDS = ['圆脸', '方脸', '菱形脸', '鹅蛋脸', '长脸', '心形脸', '梨形脸', '椭圆脸', '瓜子脸', '国字脸']
ALL_FACES = ['圆脸', '方脸', '菱形脸', '鹅蛋脸', '长脸']


def parse_faces(t):
    if not t:
        return []
    found = [w for w in FACE_WORDS if w in t]
    if found:
        return found
    if re.search(r'全部|所有|任意|都适配|都', t):
        return ALL_FACES[:]
    return []


def split_features(t):
    if not t:
        return []
    out = []
    for p in re.split(r'[、,，;；。\n]+', t):
        p = p.strip().strip('、，。；； ')
        if p:
            out.append(p)
    return out


def diff_level(d):
    if not d:
        return ('未知', 'unknown')
    t = d.strip()
    if '简单' in t or '易' in t or '低' in t:
        return (t, 'easy')
    if '难' in t or '复杂' in t or '高' in t:
        return (t, 'hard')
    return (t, 'medium')


def infer_cat(name, feat):
    txt = (name or '') + ' ' + (feat or '')
    if '长发' in txt:
        return ['long']
    if re.search(r'短发|波波|bob|精灵|超短|初恋|寸头', txt, re.I):
        return ['short']
    if re.search(r'中长|及肩|锁骨|齐肩|lob', txt, re.I):
        return ['medium']
    return ['medium']


rows = []
for r in range(2, ws.max_row + 1):
    num = ws.cell(r, 1).value
    pic = ws.cell(r, 2).value or ''
    name = ws.cell(r, 3).value
    if not name:
        continue
    intro = ws.cell(r, 4).value
    source = ws.cell(r, 5).value
    faces_txt = ws.cell(r, 6).value
    people = ws.cell(r, 7).value
    feat_txt = ws.cell(r, 8).value
    diff = ws.cell(r, 9).value
    idm = re.search(r'ID_[0-9A-Fa-f]{32}', pic)
    media = id2media.get(idm.group(0)) if idm else None
    faces = parse_faces(faces_txt)
    feats = split_features(feat_txt)
    dlabel, dlevel = diff_level(diff)
    cat = infer_cat(name, feat_txt)
    imgfile = None
    if media:
        fn = f"h{int(num):02d}.jpeg"
        with open(os.path.join(IMG_DIR, fn), 'wb') as f:
            f.write(z.read('xl/' + media))
        imgfile = f"/public/images/hair/{fn}"
    rows.append({
        "id": f"h{int(num):02d}",
        "name": name,
        "description": intro or "",
        "source": source or "",
        "suitableFaceShapes": faces,
        "suitablePeople": people or "",
        "features": feats,
        "difficulty": dlabel,
        "difficultyLevel": dlevel,
        "categories": cat,
        "imageUrl": imgfile,
        "modelUrl": None,
        "modelWidth": 1.08,
        "modelOffset": {"x": 0, "y": 0.06, "z": 0},
        "modelRotY": 0,
        "fit": "all",
        "params": {"sideLen": 0.5, "frontLen": 0.6, "volume": 0.5, "curl": 0.3},
    })

print("提取发型数:", len(rows), "| 图片保存数:", sum(1 for x in rows if x['imageUrl']))
print("样本 imageUrl:", rows[0]['imageUrl'], "| faces:", rows[0]['suitableFaceShapes'])

# 3) merge / backfill into hairDatabase.json
db_path = os.path.join(PROJ, "public/models/hair/hairDatabase.json")
db = json.load(open(db_path, encoding='utf-8'))
existing_idx = {m['id']: i for i, m in enumerate(db['models'])}
added = 0
for x in rows:
    if x['id'] in existing_idx:
        db['models'][existing_idx[x['id']]]['imageUrl'] = x['imageUrl']
    else:
        db['models'].append(x)
        added += 1
json.dump(db, open(db_path, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
open(db_path, 'a').write('\n')
print("合并后 models:", len(db['models']), "(新增", added, ", 回填图片", sum(1 for x in rows if x['imageUrl']), ")")
